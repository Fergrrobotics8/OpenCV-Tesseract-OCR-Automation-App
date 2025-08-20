import argparse, os, re, time, json, pandas as pd, xml.etree.ElementTree as ET
from pathlib import Path
from pywinauto.application import Application
from pywinauto import Desktop
import pyautogui, pytesseract, cv2, numpy as np, psutil, sys
import math
import unicodedata
import time 
from datetime import datetime
import openpyxl
import shutil

# Función auxiliar para medir tiempos (SILENCIOSA)
def log_time(start_time, operation, row_idx):
    # Tiempos desactivados para versión conservadora
    return time.time()

pyautogui.FAILSAFE = False
pyautogui.PAUSE = 0.03  # VUELTO al valor original

DECIMAL_COMMA = True
STATION_ORDER = ["RIGHTSIDE1","RIGHTSIDE2","LEFTSIDE1","LEFTSIDE2","ROOF","ROOFHOOD","HOOD","BRIDGE"]

WIN_RECT = None  # (left, top, right, bottom) cache global

def to_float(s):
    try:
        if isinstance(s, float): return s
        if isinstance(s, int): return float(s)
        s = str(s).replace(",", ".")
        return float(s)
    except Exception:
        return None

def norm_name(n):
    return str(n).strip().replace(" ","").replace("_","").upper()

def pick_col(df, preferred):
    for p in preferred:
        for c in df.columns:
            if norm_name(p) == norm_name(c):
                return c
    return None

def detect_header_row(raw, tokens=None, max_scan=60):
    tokens = tokens or ["STATION","CAMERA","FRAME","X","Y"]
    for i, row in enumerate(raw[:max_scan]):
        if any(any(t in str(cell).upper() for t in tokens) for cell in row):
            return i
    return 0

def load_excel_any_header(path, sheet_name="Hoja1"):
    #print("[INFO] Si el archivo Excel esta ABIERTO, los valores de las formulas NO se pueden leer correctamente. Por favor, CIERRA el Excel antes de ejecutar este script.")
    try:
        # Lee todo el Excel sin header
        df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)

        # Detecta la fila de headers (busca "Nr." y "X_PAG" en la fila)
        header_row_idx = None
        for i in range(df_raw.shape[0]):
            row = [str(x).strip().upper() for x in df_raw.iloc[i]]
            if "NR." in row and "X_PAG" in row:
                header_row_idx = i
                break

        if header_row_idx is None:
            print("[ERROR] No se encontró la fila de headers.")
            return pd.DataFrame(), sheet_name, 0

        # Construye el DataFrame con los headers correctos y los datos debajo
        headers = df_raw.iloc[header_row_idx]
        data = df_raw.iloc[header_row_idx+1:].reset_index(drop=True)
        data.columns = headers

        # Elimina filas completamente vacías
        data = data.dropna(how='all').reset_index(drop=True)

        return data, sheet_name, header_row_idx

    except Exception as e:
        print(f"[ERROR] Error leyendo Excel: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), sheet_name, 0
def load_stations_map(xml_path):
    stations = {}
    tree = ET.parse(xml_path)
    for st in tree.getroot().findall("station"):
        try:
            st_id = int(st.attrib["id"])
            st_name = st.findtext("name")
            stations[st_id] = st_name
        except Exception:
            continue
    return stations

def screenshot_region(rect):
    x, y, w, h = rect
    x, y, w, h = int(x), int(y), int(w), int(h)
    img = pyautogui.screenshot(region=(x, y, w, h))
    return cv2.cvtColor(np.array(img), cv2.COLOR_RGBA2BGR)

def ocr_data(img_bgr, psm=6, lang="eng+spa", timeout=2.0):
    try:
        from pytesseract import image_to_data, Output
        return pytesseract.image_to_data(img_bgr, lang=lang, config=f"--psm {psm}", output_type=Output.DICT, timeout=timeout)
    except Exception:
        return None

def merge_adjacent_words(dets, max_gap_px=18, min_overlap=0.5):
    if not dets: return []
    dets = sorted(dets, key=lambda d: (d["y"], d["x"]))
    merged = []
    skip = set()
    for i, d in enumerate(dets):
        if i in skip: continue
        group = [d]
        for j in range(i+1, len(dets)):
            dj = dets[j]
            if abs(d["y"] - dj["y"]) < max(d["h"], dj["h"])*min_overlap and 0 < dj["x"] - (group[-1]["x"] + group[-1]["w"]) < max_gap_px:
                group.append(dj)
                skip.add(j)
            else:
                break
        if len(group) == 1:
            merged.append(d)
        else:
            x = group[0]["x"]
            y = min(g["y"] for g in group)
            w = group[-1]["x"] + group[-1]["w"] - x
            h = max(g["y"]+g["h"] for g in group) - y
            text = " ".join(g["text"] for g in group)
            conf = min(g["conf"] for g in group)
            merged.append({"x":x, "y":y, "w":w, "h":h, "text":text, "conf":conf})
    return merged

def find_footer_labels_ocr(win_rect, scale=2.0, min_conf=60):
    """OCR avanzado del footer: devuelve dict {label: (x, y, w, h)} en coords de pantalla."""
    left, top, right, bottom = win_rect
    img = screenshot_region((int(left), int(top), int(right-left), int(bottom-top)))
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    data = ocr_data(gray, psm=6)
    mesh_y = None
    if data:
        for i, txt in enumerate(data["text"]):
            t = (txt or "").strip().upper()
            if "MESH" in t:
                mesh_y = int(data["top"][i] + data["height"][i]//2)
                break
    if mesh_y is None:
        mesh_y = int(gray.shape[0]*0.7)
    search_top = max(mesh_y - 120, 0)
    col = np.mean(gray[search_top:mesh_y], axis=1)
    diff = np.abs(np.diff(col))
    if len(diff) > 0:
        sep_offset = np.argmax(diff[::-1])
        sep_y = mesh_y - sep_offset
    else:
        sep_y = mesh_y
    sep_y_abs = int(top + sep_y)
    footer_img = screenshot_region((int(left), int(sep_y_abs), int(right-left), int(bottom-sep_y_abs)))
    cv2.imwrite("debug_footer_img_for_ocr.png", footer_img)
    gray_footer = cv2.cvtColor(footer_img, cv2.COLOR_BGR2GRAY)
    img_scaled = cv2.resize(gray_footer, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    data = ocr_data(img_scaled, psm=11, lang="eng+spa", timeout=2.0)
    dets = []
    if data:
        for i in range(len(data["text"])):
            txt = (data["text"][i] or "").strip()
            try: conf = float(data["conf"][i])
            except: conf = -1.0
            if txt and conf >= min_conf:
                dets.append({"x":int(data["left"][i]), "y":int(data["top"][i]), "w":int(data["width"][i]), "h":int(data["height"][i]), "text":txt, "conf":conf})
    dets_merged = merge_adjacent_words(dets, max_gap_px=18, min_overlap=0.5)
    label_targets = {
        "MESH": ["MESH"],
        "STATION": ["STATION"],
        "CAMERA": ["CAMERA"],
        "FRAME": ["FRAME"],
        "POSITION X": ["POSITION X", "X POSITION"],
        "POSITION Y": ["POSITION Y", "Y POSITION"],
        "SHOW": ["SHOW"]
    }
    labels = {}
    dbg_labels = cv2.cvtColor(img_scaled, cv2.COLOR_GRAY2BGR)
    for d in dets_merged:
        t = d["text"].strip().upper()
        for key, variants in label_targets.items():
            for v in variants:
                if t == v or t.replace(" ", "") == v.replace(" ", ""):
                    x, y, w, h = int(d["x"] / scale), int(d["y"] / scale), int(d["w"] / scale), int(d["h"] / scale)
                    labels[key] = (int(left)+x, int(sep_y_abs)+y, w, h)
                    cv2.rectangle(dbg_labels, (d["x"], d["y"]), (d["x"]+d["w"], d["y"]+d["h"]), (0, 0, 255), 2)
                    cv2.putText(dbg_labels, key, (d["x"], d["y"]-5), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,0,255), 2, cv2.LINE_AA)
    cv2.imwrite("debug_footer_labels_texts_v7.png", dbg_labels)
    thr = cv2.adaptiveThreshold(gray_footer,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,31,7)
    boxes = []
    contours, _ = cv2.findContours(255-thr, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        if w > 60 and h > 18:
            boxes.append((int(x), int(y), int(w), int(h)))
    label_boxes = {}
    for key, (lx, ly, lw, lh) in labels.items():
        candidates = [b for b in boxes if b[0] > lx+lw and abs(b[1]-ly)<20]
        if candidates:
            bx = min(candidates, key=lambda b: b[0])
            cx = bx[0] + bx[2]//2
            cy = bx[1] + bx[3]//2
            label_boxes[key] = (int(left)+cx, int(sep_y_abs)+cy, bx)
    dbg2 = footer_img.copy()
    for key, (lx, ly, lw, lh) in labels.items():
        cv2.rectangle(dbg2, (lx, ly), (lx+lw, ly+lh), (255,0,0), 1)
        cv2.putText(dbg2, key, (lx, ly-2), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255,0,0), 1, cv2.LINE_AA)
    for i, (x, y, w, h) in enumerate(boxes):
        cv2.rectangle(dbg2, (x, y), (x+w, y+h), (0,255,0), 1)
        cv2.putText(dbg2, str(i), (x+3, y+15), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,128,0), 2, cv2.LINE_AA)
    cv2.imwrite("debug_footer_labels_boxes_v7.png", dbg2)
    box_index_map = {
        "MESH": 12,
        "STATION": 8,
        "CAMERA": 5,
        "FRAME": 11,
        "POSITION X": 7,
        "POSITION Y": 4,
    }
    click_points = {}
    for key, (lx, ly, lw, lh) in labels.items():
        if key == "SHOW":
            cx = lx + lw // 2
            cy = ly + lh // 2
            click_points[key] = (cx, cy)
        elif key in box_index_map and box_index_map[key] < len(boxes):
            bx, by, bw, bh = boxes[box_index_map[key]]
            cx = int(win_rect[0]) + bx + bw // 2
            cy = int(sep_y_abs) + by + bh // 2
            click_points[key] = (cx, cy)
        else:
            click_points[key] = (lx+lw//2, ly+lh//2)
    dbg3 = footer_img.copy()
    for key, (lx, ly, lw, lh) in labels.items():
        cv2.rectangle(dbg3, (lx, ly), (lx+lw, ly+lh), (255,0,0), 1)
        cv2.putText(dbg3, key, (lx, ly-2), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255,0,0), 1, cv2.LINE_AA)
    for i, (x, y, w, h) in enumerate(boxes):
        cv2.rectangle(dbg3, (x, y), (x+w, y+h), (0,255,0), 1)
        cv2.putText(dbg3, str(i), (x+3, y+15), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,128,0), 2, cv2.LINE_AA)
    for key, (cx, cy) in click_points.items():
        cx_rel = cx - int(win_rect[0])
        cy_rel = cy - int(sep_y_abs)
        cv2.line(dbg3, (cx_rel-8, cy_rel-8), (cx_rel+8, cy_rel+8), (0,0,255), 2)
        cv2.line(dbg3, (cx_rel-8, cy_rel+8), (cx_rel+8, cy_rel-8), (0,0,255), 2)
        cv2.putText(dbg3, key, (cx_rel+10, cy_rel), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2, cv2.LINE_AA)
    cv2.imwrite("debug_footer_click_points_v7.png", dbg3)
    return click_points, (int(left), int(sep_y_abs), int(right-left), int(bottom-sep_y_abs)), footer_img, boxes

def attach_or_launch(exe_path, title_hint, work_dir=None, attach_only=False, extra_load_wait=1.0):
    exe_basename = os.path.basename(exe_path)
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] and exe_basename.lower() in proc.info['name'].lower():
            app = Application(backend="uia").connect(process=proc.info['pid'])
            break
    else:
        if attach_only:
            raise RuntimeError("No se encontro el proceso y attach_only=True")
        app = Application(backend="uia").start(exe_path, work_dir=work_dir)
        time.sleep(extra_load_wait)
    windows = Desktop(backend="uia").windows()
    for w in windows:
        if title_hint.lower() in w.window_text().lower():
            w.set_focus()
            w.maximize()
            rect = w.rectangle()
            global WIN_RECT
            WIN_RECT = (rect.left, rect.top, rect.right, rect.bottom)
            return w
    raise RuntimeError("No se encontro la ventana con el titulo esperado")

# OCR robusto SILENCIOSO (sin debug images)
def ocr_single_digit_robust(img, target_digit, row_idx):
    """OCR especializado para detectar un digito especifico en dropdown de camaras."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) > 2 else img.copy()
    
    # Configuraciones de transformacion para mejorar deteccion de digitos finos
    transforms = []
    
    # 1. Zoom agresivo + threshold adaptativo
    zoom = 4.0
    zoomed = cv2.resize(gray, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
    transforms.append(("zoom4x", zoomed))
    
    # 2. Otsu threshold con zoom
    _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    otsu_zoom = cv2.resize(otsu, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
    transforms.append(("otsu_zoom", otsu_zoom))
    
    # 3. Otsu invertido (texto blanco sobre fondo negro)
    otsu_inv_zoom = cv2.resize(255 - otsu, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
    transforms.append(("otsu_inv_zoom", otsu_inv_zoom))
    
    # 4. Threshold adaptativo con zoom
    adaptive = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    adaptive_zoom = cv2.resize(adaptive, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
    transforms.append(("adaptive_zoom", adaptive_zoom))
    
    results = []
    target_str = str(target_digit)
    
    for i, (name, img_transform) in enumerate(transforms):
        # Configuracion especifica para digitos solamente
        config = "--psm 6 --oem 1 -c tessedit_char_whitelist=0123456789"
        
        try:
            data = pytesseract.image_to_data(img_transform, lang="eng", config=config, output_type=pytesseract.Output.DICT)
            
            for j, text in enumerate(data["text"]):
                text = text.strip()
                if not text:
                    continue
                    
                conf = float(data["conf"][j]) if str(data["conf"][j]).replace('.','',1).isdigit() else 0.0
                
                if conf >= 30:  # Bajamos el umbral para digitos dificiles
                    try:
                        detected_digit = int(text)
                        if detected_digit == target_digit:
                            # Coordenadas ajustadas por el zoom
                            x = int(data["left"][j] / zoom)
                            y = int(data["top"][j] / zoom)
                            w = int(data["width"][j] / zoom)
                            h = int(data["height"][j] / zoom)
                            cx, cy = x + w//2, y + h//2
                            
                            results.append({
                                "center": (cx, cy),
                                "conf": conf,
                                "transform": name,
                                "bbox": (x, y, w, h)
                            })
                    except ValueError:
                        continue
        except Exception as e:
            continue
    
    # SIN generar imagen debug - solo retorna el resultado
    if results:
        best = max(results, key=lambda x: x["conf"])
        print(f"[DEBUG] Camara {target_digit} encontrada con metodo robusto!")
        return best
    else:
        print(f"[WARN] Camara {target_digit} no encontrada con metodo robusto")
        return None

intersect_table = []

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="config_v6_5e.json")
    ap.add_argument("--run", action="store_true")
    ap.add_argument("--dry", action="store_true")
    args = ap.parse_args()

    cfg = json.load(open(args.config, "r", encoding="utf-8"))
    if cfg.get("tesseract_cmd"):
        pytesseract.pytesseract.tesseract_cmd = cfg["tesseract_cmd"]

    df, sheet, hdr = load_excel_any_header(cfg["excel_path"], sheet_name=cfg.get("excel_sheet", "Hoja1"))

    col_station = pick_col(df, ["STATION", "STATIONID", "ESTACION"])
    col_camera = pick_col(df, ["CAMERA", "CAM", "CAMERAID"])
    col_frame = pick_col(df, ["FRAME"])
    col_xf = pick_col(df, ["X_FRAME", "XFRAME", "XFRM"])
    col_yf = pick_col(df, ["Y_FRAME", "YFRAME", "YFRM"])
    col_nr = pick_col(df, ["NR", "Nr.", "Nr", "NUMERO", "NuMERO", "NUMBER"])

    if args.dry:
        print("[DRY] Header row index:", hdr)
        print("[DRY] Columns:", list(df.columns))
        print("[DRY] Mapping -> STATION:", col_station, "CAMERA:", col_camera, "FRAME:", col_frame, "X_FRAME:", col_xf, "Y_FRAME:", col_yf, "NR:", col_nr)
        print("[DRY] Head:\n", df.head(3))
        return

    if not all([col_station, col_camera, col_frame, col_xf, col_yf]):
        print("[ERR] No encuentro columnas. Tengo:", list(df.columns))
        return

    stations = load_stations_map(cfg["stations_xml_path"])

    # Launch viewer
    dlg = attach_or_launch(
        cfg["viewer_exe_path"],
        cfg.get("window_title_contains", "EINES System 3D for ESPQi"),
        work_dir=cfg.get("work_dir"),
        attach_only=bool(cfg.get("attach_only", False)),
        extra_load_wait=float(cfg.get("extra_load_wait_s", 3.0))
    )

    # Deteccion robusta de footer y labels/cajas
    label_boxes, (footer_left, footer_top, footer_w, footer_h), footer_img, boxes = find_footer_labels_ocr(WIN_RECT, scale=2.0, min_conf=60)

    # Selecciona Mesh solo una vez
    if "MESH" in label_boxes:
        mx, my = label_boxes["MESH"]
        pyautogui.moveTo(mx, my, duration=0.2)
        pyautogui.click()
        print("[DEBUG] Clic en la caja de MESH")
        time.sleep(1.0)  # Espera a que se abra el desplegable

        # Usar la misma región que el dropdown de camaras para el OCR de Mesh
        screen_w, screen_h = pyautogui.size()
        footer_bottom = int(footer_top + footer_h)
        taskbar_height = screen_h - footer_bottom
        if taskbar_height < 30 or taskbar_height > 80:
            taskbar_height = 40  # fallback si el calculo es raro
        top = screen_h // 2
        height = (screen_h - taskbar_height) - top

        dropdown_region = (footer_left, top, footer_w, height)
        dropdown_img = screenshot_region(dropdown_region)
        cv2.imwrite("debug_dropdown_mesh.png", dropdown_img)

        # OCR sobre el desplegable
        gray_dd = cv2.cvtColor(dropdown_img, cv2.COLOR_BGR2GRAY)
        data_dd = ocr_data(gray_dd, psm=11, lang="eng+spa", timeout=2.0)
        found = False

        # Debug: mostrar todos los textos detectados en el dropdown
        dbg_mesh = dropdown_img.copy()
        print("[DEBUG] Textos detectados en dropdown Mesh:")
        if data_dd:
            for i, txt in enumerate(data_dd["text"]):
                t = (txt or "").strip()
                conf = float(data_dd["conf"][i]) if str(data_dd["conf"][i]).replace('.','',1).isdigit() else -1.0
                x, y, w, h = int(data_dd["left"][i]), int(data_dd["top"][i]), int(data_dd["width"][i]), int(data_dd["height"][i])
                if t:
                    print(f"  [{i}] '{t}' (conf: {conf:.1f}) pos: ({x},{y},{w},{h})")
                    color = (0,255,0)
                    if "TAYCAN_NACHFOLGER" in t.upper():
                        color = (0,0,255)
                    cv2.rectangle(dbg_mesh, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(dbg_mesh, t, (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2, cv2.LINE_AA)
                if "TAYCAN_NACHFOLGER" in t.upper():
                    cx = dropdown_region[0] + x + w//2
                    cy = dropdown_region[1] + y + h//2
                    pyautogui.moveTo(cx, cy, duration=0.1)
                    pyautogui.click()
                    found = True
                    print("[DEBUG] Seleccionado TAYCAN_NACHFOLGER en Mesh")
                    # Dibuja rectangulo especial y guarda imagen de debug
                    cv2.rectangle(dbg_mesh, (x, y), (x+w, y+h), (0,0,255), 3)
                    break
            cv2.imwrite("debug_dropdown_mesh_all_detected.png", dbg_mesh)
        else:
            print("[WARN] OCR no devolvió resultados para el dropdown Mesh.")

        if not found:
            print("[WARN] No se encontro TAYCAN_NACHFOLGER en el desplegable de Mesh")
        time.sleep(1.0)

    # CACHE GLOBAL para estaciones (mantener optimización)
    station_dropdown_boxes = None
    
        # --- Mapeo de fila Excel <-> NR para escritura robusta ---
    excel_path = cfg["excel_path"]
    # Comprobar si el archivo está abierto (por lock)
    excel_open = False
    try:
        with open(excel_path, "a"):
            pass
    except PermissionError:
        excel_open = True
    
    if excel_open:
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_copy = excel_path.replace(".xlsx", f"_copia_{now_str}.xlsx")
        shutil.copy2(excel_path, excel_copy)
        print(f"[INFO] El archivo Excel original esta ABIERTO. Se ha creado una copia: {excel_copy} y se escribira en esa copia.")
        target_excel = excel_copy
    else:
        target_excel = excel_path
    
    # Leer SOLO UNA VEZ los valores de NR y la fila real
    wb_map = openpyxl.load_workbook(target_excel, data_only=True)
    sheet_name = cfg.get("excel_sheet", "Hoja1")
    ws_map = wb_map[sheet_name]
    header_row = hdr + 1  # hdr viene de load_excel_any_header, openpyxl es 1-based
    
    # Buscar la columna NR
    col_nr_excel = None
    for col in range(1, ws_map.max_column + 1):
        cell_val = ws_map.cell(row=header_row, column=col).value
        if cell_val and str(cell_val).strip().upper() in ["NR", "NR.", "NUMERO", "NUMBER"]:
            col_nr_excel = col
            break
    if not col_nr_excel:
        print(f"[ERROR] No se encontró la columna NR en la hoja {sheet_name}. No se escribirá nada en el Excel.")
        wb_map.close()
        nr_to_row = {}
    else:
        # Mapeo NR -> fila
        nr_to_row = {}
        for rowi in range(header_row + 1, ws_map.max_row + 1):
            nr_val = ws_map.cell(row=rowi, column=col_nr_excel).value
            if nr_val is not None:
                nr_to_row[str(nr_val).strip()] = rowi
        print(f"[DEBUG] Mapeo NR->fila: {nr_to_row}")
        wb_map.close()

    for idx, row in df.iterrows():
        try:
            # Conversión segura de valores
            def safe_int(val, default="NA"):
                try:
                    if pd.isna(val): return default
                    return int(float(val))
                except Exception:
                    return default

            st_name = str(row[col_station])
            station_id = safe_int(row[col_station])
            camera_id = safe_int(row[col_camera])
            frame = str(safe_int(row[col_frame]))
            x_val = str(safe_int(row[col_xf]))
            y_val = str(safe_int(row[col_yf]))

            # Si algún valor es "NA", saltar la fila
            if "NA" in [station_id, camera_id, frame, x_val, y_val]:
                print(f"[WARN] Row {idx}: Datos incompletos, saltando fila. Valores: station={station_id}, camera={camera_id}, frame={frame}, x={x_val}, y={y_val}")
                continue

            found_station = False
            found_camera = False

            # --- Selecciona estacion ---
            station_name = stations.get(station_id, str(station_id)).upper()
            if "STATION" in label_boxes:
                sx, sy = label_boxes["STATION"]
                
                # Clic en station
                pyautogui.moveTo(sx, sy, duration=0.1)
                pyautogui.click()
                
                # MEJORA: Espera dropdown REDUCIDA
                time.sleep(0.4)  # Reducido de 0.5 a 0.4

                # Calculo de region OCR
                if len(boxes) > 8:
                    bx, by, bw, bh = boxes[8]
                    left = bx + footer_left
                    right = bx + bw + footer_left
                else:
                    left = 0
                    right = 300  # fallback

                screen_w, screen_h = pyautogui.size()
                footer_bottom = int(footer_top + footer_h)
                taskbar_height = screen_h - footer_bottom
                if taskbar_height < 30 or taskbar_height > 80:
                    taskbar_height = 40  # fallback si el calculo es raro
                top = screen_h // 2
                height = (screen_h - taskbar_height) - top

                dropdown_region = (footer_left, top, footer_w, height)
                dropdown_img = screenshot_region(dropdown_region)
                cv2.imwrite("debug_dropdown_mesh.png", dropdown_img)

                # OCR procesamiento
                zoom = 2.0
                zoomed_img = cv2.resize(dropdown_img, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
                gray_station_dd = cv2.cvtColor(zoomed_img, cv2.COLOR_BGR2GRAY)
                data_station_dd = ocr_data(gray_station_dd, psm=11, lang="eng+spa", timeout=2.0)

                # OPTIMIZACIÓN: Procesamiento de bounding boxes (solo primera vez)
                if station_dropdown_boxes is None:
                    station_dropdown_boxes = []
                    dbg_station = zoomed_img.copy()
                    if data_station_dd:
                        for i, txt in enumerate(data_station_dd["text"]):
                            t = (txt or "").strip().upper()
                            if not t: continue
                            x, y, w, h = int(data_station_dd["left"][i]), int(data_station_dd["top"][i]), int(data_station_dd["width"][i]), int(data_station_dd["height"][i])
                            cx = x + w // 2
                            cy = y + h // 2
                            station_dropdown_boxes.append({
                                "text": t,
                                "bbox": (x, y, w, h),
                                "center": (cx, cy)
                            })
                            cv2.rectangle(dbg_station, (x, y), (x+w, y+h), (0,0,255), 2)
                            cv2.putText(dbg_station, t, (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2, cv2.LINE_AA)
                        cv2.line(dbg_station, (0, height-1), (dbg_station.shape[1], height-1), (0,0,255), 2)
                    cv2.imwrite("debug_dropdown_station_all_detected.png", dbg_station)

                # Debug visual para fila actual
                print(f"[DEBUG] Row {idx}: station_id={station_id}, station_name={station_name}, camera_id={camera_id}, frame={frame}, x={x_val}, y={y_val}")
                found_station = False
                if station_dropdown_boxes:
                    for box in station_dropdown_boxes:
                        box_text = box["text"].strip().upper()
                        if box_text == station_name:
                            cx, cy = box["center"]
                            cx_abs = int(left) + int(cx // zoom)
                            cy_abs = int(top) + int(cy // zoom)
                            pyautogui.moveTo(cx_abs, cy_abs, duration=0.1)
                            pyautogui.click()
                            found_station = True
                            print(f"[DEBUG] Seleccionada estacion: {station_name}")
                            break
                
                # Solo genera debug visual si no encontro la estacion
                if not found_station:
                    dbg_station_click = dbg_station.copy()
                    cv2.putText(dbg_station_click, f"NO EXACT MATCH: {station_name}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2, cv2.LINE_AA)
                    for i, box in enumerate(station_dropdown_boxes[:5]):
                        cv2.putText(dbg_station_click, f"Available: {box['text']}", (10, 60 + i*25), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,0,0), 1, cv2.LINE_AA)
                    cv2.imwrite(f"debug_dropdown_station_centers_row{idx}.png", dbg_station_click)
                    print(f"[WARN] Row {idx}: No se encontro {station_name} en el desplegable de stations")
                
                # MEJORA: Espera post-clic REDUCIDA
                time.sleep(0.2)  # Reducido de 0.3 a 0.2

            # --- Selecciona camara ---
            if "CAMERA" in label_boxes:
                cx, cy = label_boxes["CAMERA"]
                
                # Clic en camera
                pyautogui.moveTo(cx, cy, duration=0.1)
                pyautogui.click()
                
                # MEJORA: Espera dropdown REDUCIDA
                time.sleep(0.4)  # Reducido de 0.5 a 0.4

                # Screenshot
                camera_dropdown_region = (left, top, right - left, height)
                camera_dropdown_img = screenshot_region(camera_dropdown_region)

                # OCR
                zoom = 2.0
                zoomed_img = cv2.resize(camera_dropdown_img, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
                gray_camera_dd = cv2.cvtColor(zoomed_img, cv2.COLOR_BGR2GRAY)
                data_camera_dd = ocr_data(gray_camera_dd, psm=11, lang="eng+spa", timeout=2.0)

                # Procesamiento bboxes y busqueda
                camera_dropdown_boxes = []
                if data_camera_dd:
                    for i, txt in enumerate(data_camera_dd["text"]):
                        t = (txt or "").strip().upper()
                        if not t: continue
                        x, y, w, h = int(data_camera_dd["left"][i]), int(data_camera_dd["top"][i]), int(data_camera_dd["width"][i]), int(data_camera_dd["height"][i])
                        cx_box = x + w // 2
                        cy_box = y + h // 2
                        camera_dropdown_boxes.append({
                            "text": t,
                            "bbox": (x, y, w, h),
                            "center": (cx_box, cy_box)
                        })

                try:
                    camera_id = str(row[col_camera]).strip()
                    camera_id_int = int(camera_id)
                except Exception as e:
                    print(f"[ERR] Row {idx}: camera_id mal formado: {row[col_camera]} ({e})")
                    continue
                
                found_camera = False
                cam_str = str(camera_id_int)
                for box in camera_dropdown_boxes:
                    t = box["text"].upper().strip()
                    if re.search(rf"(?:^|[^0-9])0*{cam_str}(?:$|[^0-9])", t):
                        cx, cy = box["center"]
                        cx_abs = int(left) + int(cx // zoom)
                        cy_abs = int(top) + int(cy // zoom)
                        pyautogui.moveTo(cx_abs, cy_abs, duration=0.1)
                        pyautogui.click()
                        found_camera = True
                        print(f"[DEBUG] Seleccionada camara: {camera_id_int}")
                        break
                
                if not found_camera:
                    print(f"[DEBUG] Row {idx}: Primera pasada fallo. Intentando metodo robusto para camara {camera_id_int}")
                    robust_result = ocr_single_digit_robust(camera_dropdown_img, camera_id_int, idx)
                    
                    if robust_result:
                        cx_rel, cy_rel = robust_result["center"]
                        cx_abs = int(camera_dropdown_region[0] + cx_rel)
                        cy_abs = int(camera_dropdown_region[1] + cy_rel)
                        
                        pyautogui.moveTo(cx_abs, cy_abs, duration=0.1)
                        pyautogui.click()
                        found_camera = True
                    else:
                        print(f"[WARN] Row {idx}: Camara {camera_id_int} no encontrada ni con metodo robusto")
                
                if not found_camera:
                    print(f"[WARN] Row {idx}: No se encontro camara {camera_id_int} en el desplegable de camaras")
                    # Solo genera debug si no encontro
                    dbg_camera = zoomed_img.copy()
                    for i, box in enumerate(camera_dropdown_boxes):
                        x, y, w, h = box["bbox"]
                        cx_box, cy_box = box["center"]
                        cv2.circle(dbg_camera, (cx_box, cy_box), 7, (0,0,255), -1)
                        cv2.putText(dbg_camera, box["text"], (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2, cv2.LINE_AA)
                    cv2.line(dbg_camera, (0, height-1), (dbg_camera.shape[1], height-1), (0,0,255), 2)
                    cv2.imwrite(f"debug_dropdown_camera_centers_row{idx}.png", dbg_camera)
                
                # MEJORA: Espera post-clic REDUCIDA
                time.sleep(0.2)  # Reducido de 0.3 a 0.2

            # --- Frame ---
            if "FRAME" in label_boxes:
                fx, fy = label_boxes["FRAME"]
                
                pyautogui.moveTo(fx, fy, duration=0.1)
                pyautogui.click()
                
                pyautogui.keyDown('ctrl')
                pyautogui.press('a')  # Seleccionar todo
                pyautogui.keyUp('ctrl')
                pyautogui.press('delete')  # Borrar seleccion
                
                # MEJORA: Espera REDUCIDA
                time.sleep(0.15)  # Reducido de 0.2 a 0.15
                
                pyautogui.typewrite(frame)

            # --- Position X ---
            if "POSITION X" in label_boxes:
                px, py_ = label_boxes["POSITION X"]
                
                pyautogui.moveTo(px, py_, duration=0.1)
                pyautogui.click()
                
                pyautogui.keyDown('ctrl')
                pyautogui.press('a')  # Seleccionar todo
                pyautogui.keyUp('ctrl')
                pyautogui.press('delete')  # Borrar seleccion
                
                # MEJORA: Espera REDUCIDA
                time.sleep(0.15)  # Reducido de 0.2 a 0.15
                
                pyautogui.typewrite(x_val)

            # --- Position Y ---
            if "POSITION Y" in label_boxes:
                pyx, pyy = label_boxes["POSITION Y"]
                
                pyautogui.moveTo(pyx, pyy, duration=0.1)
                pyautogui.click()
                
                pyautogui.keyDown('ctrl')
                pyautogui.press('a')  # Seleccionar todo
                pyautogui.keyUp('ctrl')
                pyautogui.press('delete')  # Borrar seleccion
                
                # MEJORA: Espera REDUCIDA
                time.sleep(0.15)  # Reducido de 0.2 a 0.15
                
                pyautogui.typewrite(y_val)

            # --- Click SHOW ---
            if "SHOW" in label_boxes:
                shx, shy = label_boxes["SHOW"]
                
                pyautogui.moveTo(shx, shy, duration=0.1)
                pyautogui.click()
                print("[DEBUG] Click en SHOW")
                
                # Mantener 1 segundo tras SHOW
                time.sleep(1.0)

                # --- OCR del footer ---
                footer_img_live = screenshot_region((int(footer_left), int(footer_top), int(footer_w), int(footer_h)))
                zoom = 2.0
                gray_footer = cv2.cvtColor(footer_img_live, cv2.COLOR_BGR2GRAY)
                zoomed_footer = cv2.resize(gray_footer, None, fx=zoom, fy=zoom, interpolation=cv2.INTER_CUBIC)
                data_footer = ocr_data(zoomed_footer, psm=11, lang="eng+spa", timeout=2.0)

                # Procesamiento footer
                dets = []
                if data_footer:
                    for i in range(len(data_footer["text"])):
                        txt = (data_footer["text"][i] or "").strip()
                        try: conf = float(data_footer["conf"][i])
                        except: conf = -1.0
                        if txt and conf >= 60:
                            dets.append({
                                "x": int(data_footer["left"][i]),
                                "y": int(data_footer["top"][i]),
                                "w": int(data_footer["width"][i]),
                                "h": int(data_footer["height"][i]),
                                "text": txt,
                                "conf": conf
                            })
                dets_merged = merge_adjacent_words(dets, max_gap_px=18, min_overlap=0.5)

                # Busqueda intersect CON MÚLTIPLES PASADAS ROBUSTAS + DEBUG IMAGES
                found_intersect = False
                xyz = ["NA", "NA", "NA"]
                
                # GUARDA imagen base del footer para debug
                cv2.imwrite(f"debug_intersect_footer_raw_row{idx}.png", footer_img_live)
                cv2.imwrite(f"debug_intersect_footer_zoomed_row{idx}.png", zoomed_footer)
                
                # PASADA 1: Búsqueda directa en texto fusionado
                print(f"[DEBUG] Row {idx}: === INICIANDO PASADA 1 (texto fusionado) ===")
                for d in dets_merged:
                    text = d["text"].strip()
                    text_upper = text.upper()
                    if "INTERSECT POSITION" in text_upper:
                        intersect_text = text
                        # Añadir espacio tras los dos puntos si falta (para que el primer número se detecte)
                        intersect_text = re.sub(r':([-\d])', r': \1', intersect_text)
                        print(f"[DEBUG] Row {idx}: ENCONTRADO 'Intersect position' (PASADA 1)")
                        print(f"[DEBUG] Row {idx}: Texto completo: '{intersect_text}'")
                        coordinate_pattern = r'-?\d{1,4},\d{1,3}'
                        coord_matches = re.findall(coordinate_pattern, intersect_text)
                        print(f"[DEBUG] Row {idx}: Coordenadas encontradas: {coord_matches}")
                        if len(coord_matches) >= 3:
                            X, Y, Z = coord_matches[0], coord_matches[1], coord_matches[2]
                            xyz = [X, Y, Z]
                            found_intersect = True
                            print(f"[DEBUG] Row {idx}: COORDENADAS EXTRAIDAS (PASADA 1)! X={X}, Y={Y}, Z={Z}")
                        else:
                            print(f"[DEBUG] Row {idx}: Solo {len(coord_matches)} coordenadas encontradas, necesito 3")
                        break

                # PASADA 2: Si no encontró, buscar en detecciones individuales SOLO en la línea de Intersect
                if not found_intersect:
                    print(f"[DEBUG] Row {idx}: === PASADA 1 FALLÓ - INICIANDO PASADA 2 (detecciones individuales) ===")
                    if data_footer:
                        for i in range(len(data_footer["text"])):
                            txt = (data_footer["text"][i] or "").strip()
                            try: conf = float(data_footer["conf"][i])
                            except: conf = -1.0
                            if txt and conf >= 40:
                                txt_upper = txt.upper()
                                if "INTERSECT" in txt_upper and ("POSITION" in txt_upper or "POS" in txt_upper):
                                    print(f"[DEBUG] Row {idx}: ENCONTRADO fragmento 'Intersect' (PASADA 2): '{txt}'")
                                    # Añadir espacio tras los dos puntos si falta
                                    txt_fixed = re.sub(r':([-\d])', r': \1', txt)
                                    coords_found = []
                                    coord_pattern = r'-?\d{1,4},\d{1,3}'
                                    coords_found.extend(re.findall(coord_pattern, txt_fixed))
                                    # También mirar la siguiente detección si está cerca en Y
                                    if i+1 < len(data_footer["text"]):
                                        next_txt = (data_footer["text"][i+1] or "").strip()
                                        next_y = int(data_footer["top"][i+1])
                                        this_y = int(data_footer["top"][i])
                                        if abs(next_y - this_y) < 40:
                                            next_txt_fixed = re.sub(r':([-\d])', r': \1', next_txt)
                                            coords_found.extend(re.findall(coord_pattern, next_txt_fixed))
                                    print(f"[DEBUG] Row {idx}: Coordenadas en Intersect position (PASADA 2): {coords_found}")
                                    if len(coords_found) >= 3:
                                        X, Y, Z = coords_found[0], coords_found[1], coords_found[2]
                                        xyz = [X, Y, Z]
                                        found_intersect = True
                                        print(f"[DEBUG] Row {idx}: COORDENADAS EXTRAIDAS (PASADA 2)! X={X}, Y={Y}, Z={Z}")
                                    else:
                                        print(f"[DEBUG] Row {idx}: Solo {len(coords_found)} coordenadas en Intersect, necesito 3")
                                    break

                # PASADA 3: OCR más agresivo, SOLO línea con Intersect
                if not found_intersect:
                    print(f"[DEBUG] Row {idx}: === PASADA 2 FALLÓ - INICIANDO PASADA 3 (OCR agresivo) ===")
                    data_footer_aggressive = ocr_data(zoomed_footer, psm=6, lang="eng+spa", timeout=3.0)
                    if data_footer_aggressive:
                        for i in range(len(data_footer_aggressive["text"])):
                            txt = (data_footer_aggressive["text"][i] or "").strip()
                            if "INTERSECT" in txt.upper():
                                txt_fixed = re.sub(r':([-\d])', r': \1', txt)
                                coord_pattern = r'-?\d{1,4},\d{1,3}'
                                coords_found = re.findall(coord_pattern, txt_fixed)
                                # También mirar la siguiente detección si está cerca en Y
                                if i+1 < len(data_footer_aggressive["text"]):
                                    next_txt = (data_footer_aggressive["text"][i+1] or "").strip()
                                    next_y = int(data_footer_aggressive["top"][i+1])
                                    this_y = int(data_footer_aggressive["top"][i])
                                    if abs(next_y - this_y) < 40:
                                        next_txt_fixed = re.sub(r':([-\d])', r': \1', next_txt)
                                        coords_found.extend(re.findall(coord_pattern, next_txt_fixed))
                                print(f"[DEBUG] Row {idx}: Coordenadas en Intersect position (PASADA 3): {coords_found}")
                                if len(coords_found) >= 3:
                                    X, Y, Z = coords_found[0], coords_found[1], coords_found[2]
                                    xyz = [X, Y, Z]
                                    found_intersect = True
                                    print(f"[DEBUG] Row {idx}: COORDENADAS EXTRAIDAS (PASADA 3)! X={X}, Y={Y}, Z={Z}")
                                else:
                                    print(f"[DEBUG] Row {idx}: Solo {len(coords_found)} coordenadas en Intersect, necesito 3")
                                break
                
                # PASADA 4: ROI ESPECÍFICA si aún no encontró nada
                if not found_intersect:
                    print(f"[DEBUG] Row {idx}: === PASADA 3 FALLÓ - INICIANDO PASADA 4 (ROI específica) ===")
                    roi_x1, roi_y1, roi_x2, roi_y2 = 1210, 110, 2120, 50
                    img_h, img_w = zoomed_footer.shape[:2]
                    roi_x1 = max(0, min(roi_x1, img_w))
                    roi_y1 = max(0, min(roi_y1, img_h))
                    roi_x2 = max(0, min(roi_x2, img_w))
                    roi_y2 = max(0, min(roi_y2, img_h))
                    if roi_y1 > roi_y2:
                        roi_y1, roi_y2 = roi_y2, roi_y1
                    if roi_x1 > roi_x2:
                        roi_x1, roi_x2 = roi_x2, roi_x1
                    print(f"[DEBUG] Row {idx}: ROI ajustada: ({roi_x1}, {roi_y1}) a ({roi_x2}, {roi_y2})")
                    print(f"[DEBUG] Row {idx}: Imagen zoomed footer size: {img_w}x{img_h}")
                    roi_img = zoomed_footer[roi_y1:roi_y2, roi_x1:roi_x2]
                    cv2.imwrite(f"debug_intersect_roi_raw_row{idx}.png", roi_img)

                    if roi_img.size > 0:
                        print(f"[DEBUG] Row {idx}: ROI extraída size: {roi_img.shape}")
                        roi_zoom = 1.5
                        roi_zoomed = cv2.resize(roi_img, None, fx=roi_zoom, fy=roi_zoom, interpolation=cv2.INTER_CUBIC)
                        cv2.imwrite(f"debug_intersect_roi_zoomed_row{idx}.png", roi_zoomed)
                        roi_configs = [
                            ("psm6", 6),
                            ("psm7", 7),
                            ("psm8", 8),
                            ("psm11", 11)
                        ]
                        roi_debug_results = cv2.cvtColor(roi_zoomed, cv2.COLOR_GRAY2BGR)
                        best_coords = None
                        best_score = -1
                        best_text = ""
                        for config_name, psm_val in roi_configs:
                            try:
                                data_roi = ocr_data(roi_zoomed, psm=psm_val, lang="eng+spa", timeout=2.0)
                                if data_roi:
                                    roi_text_parts = []
                                    for i in range(len(data_roi["text"])):
                                        txt = (data_roi["text"][i] or "").strip()
                                        try: conf = float(data_roi["conf"][i])
                                        except: conf = 0.0
                                        if txt and conf >= 30:
                                            roi_text_parts.append(txt)
                                            x = int(data_roi["left"][i])
                                            y = int(data_roi["top"][i])
                                            w = int(data_roi["width"][i])
                                            h = int(data_roi["height"][i])
                                            if re.match(r'-?\d{1,4},\d{1,3}$', txt):
                                                color = (0, 255, 0)
                                                font_scale = 1.2
                                                thickness = 2
                                                cv2.rectangle(roi_debug_results, (x, y), (x+w, y+h), color, 2)
                                                cv2.putText(roi_debug_results, txt, (x, y+h+30), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)
                                            else:
                                                color = (255, 0, 0)
                                                font_scale = 0.6
                                                thickness = 1
                                                cv2.rectangle(roi_debug_results, (x, y), (x+w, y+h), color, 1)
                                                cv2.putText(roi_debug_results, txt, (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)
                                    roi_full_text = " ".join(roi_text_parts)
                                    print(f"[DEBUG] Row {idx}: ROI {config_name} texto: '{roi_full_text}'")
                                    match = re.search(r'INTERSECT POSITION:?\s*([-\d,\. ]+)', roi_full_text.upper())
                                    coords_found = []
                                    if match:
                                        coord_text = match.group(1).split('(')[0].strip()
                                        coordinate_pattern = r'-?\d{1,4},\d{1,3}'
                                        coords_found = re.findall(coordinate_pattern, coord_text)
                                        print(f"[DEBUG] Row {idx}: Coordenadas tras Intersect position: {coords_found}")
                                        # MAGIA: reconstrucción si la Z está partida
                                        if len(coords_found) > 3:
                                            last = coords_found[-2:]
                                            if all(re.match(r'^\d{1,3}$', x.replace(',', '')) for x in last):
                                                coords_found[-2] = coords_found[-2] + coords_found[-1]
                                                coords_found = coords_found[:-1]
                                            elif re.match(r'^\d{1,3}$', coords_found[-2].replace(',', '')) and re.match(r'^,\d{1,3}$', coords_found[-1]):
                                                coords_found[-2] = coords_found[-2] + coords_found[-1]
                                                coords_found = coords_found[:-1]
                                        # Solo si hay 3 coordenadas
                                        if len(coords_found) == 3:
                                            score = sum(len(c) for c in coords_found)
                                            if score > best_score:
                                                best_score = score
                                                best_coords = coords_found
                                                best_text = roi_full_text
                            except Exception as e:
                                print(f"[DEBUG] Row {idx}: Error en ROI OCR {config_name}: {e}")
                                continue
                        cv2.imwrite(f"debug_intersect_roi_detections_row{idx}.png", roi_debug_results)
                        if best_coords:
                            X, Y, Z = best_coords[0], best_coords[1], best_coords[2]
                            xyz = [X, Y, Z]
                            found_intersect = True
                            print(f"[DEBUG] Row {idx}: COORDENADAS FINALES (PASADA 4)! X={X}, Y={Y}, Z={Z} (texto: '{best_text}')")
                        else:
                            print(f"[DEBUG] Row {idx}: ROI no produjo resultados válidos para Intersect position.")
                    else:
                        print(f"[DEBUG] Row {idx}: ROI vacía o fuera de límites")
        
                # === GENERACIÓN DE IMÁGENES DEBUG PARA INTERSECT ===
                dbg_intersect = cv2.cvtColor(zoomed_footer, cv2.COLOR_GRAY2BGR)
                
                # Dibuja TODAS las detecciones fusionadas
                for i, d in enumerate(dets_merged):
                    x, y, w, h = d["x"], d["y"], d["w"], d["h"]
                    text = d["text"]
                    conf = d["conf"]
                    
                    # Color según si contiene "INTERSECT"
                    if "INTERSECT" in text.upper():
                        color = (0, 255, 0)  # Verde si contiene INTERSECT
                        thickness = 3
                    else:
                        color = (255, 0, 0)  # Azul para otros textos
                        thickness = 1
                    
                    cv2.rectangle(dbg_intersect, (x, y), (x+w, y+h), color, thickness)
                    cv2.putText(dbg_intersect, f"{i}:{text}", (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1, cv2.LINE_AA)
                    cv2.putText(dbg_intersect, f"C:{conf:.0f}", (x, y+h+15), cv2.FONT_HERSHEY_SIMPLEX, 0.3, color, 1, cv2.LINE_AA)
                
                # ACTIVAR GUARDADO DE IMÁGENES DEBUG
                cv2.imwrite(f"debug_intersect_merged_detections_row{idx}.png", dbg_intersect)
                
                # Dibuja TODAS las detecciones individuales (si hay datos)
                if data_footer:
                    dbg_individual = cv2.cvtColor(zoomed_footer, cv2.COLOR_GRAY2BGR)
                    for i in range(len(data_footer["text"])):
                        txt = (data_footer["text"][i] or "").strip()
                        if not txt: continue
                        
                        try: conf = float(data_footer["conf"][i])
                        except: conf = -1.0
                        
                        if conf >= 20:  # Mostrar solo detecciones con confianza razonable
                            x = int(data_footer["left"][i])
                            y = int(data_footer["top"][i])
                            w = int(data_footer["width"][i])
                            h = int(data_footer["height"][i])
                            
                            # Color según contenido y confianza
                            if "INTERSECT" in txt.upper():
                                color = (0, 255, 0)  # Verde para INTERSECT
                                thickness = 3
                            elif conf >= 60:
                                color = (255, 0, 0)  # Azul para alta confianza
                                thickness = 2
                            else:
                                color = (0, 165, 255)  # Naranja para baja confianza
                                thickness = 1
                            
                            cv2.rectangle(dbg_individual, (x, y), (x+w, y+h), color, thickness)
                            cv2.putText(dbg_individual, f"{i}:{txt}", (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1, cv2.LINE_AA)
                            cv2.putText(dbg_individual, f"C:{conf:.0f}", (x, y+h+15), cv2.FONT_HERSHEY_SIMPLEX, 0.3, color, 1, cv2.LINE_AA)
                    
                    # ACTIVAR GUARDADO DE IMÁGENES DEBUG
                    cv2.imwrite(f"debug_intersect_individual_detections_row{idx}.png", dbg_individual)
                
                # Imagen de resumen con resultado final
                dbg_result = cv2.cvtColor(zoomed_footer, cv2.COLOR_GRAY2BGR)
                result_text = f"Row {idx}: {'FOUND' if found_intersect else 'NOT FOUND'}"
                coords_text = f"XYZ: {xyz[0]}, {xyz[1]}, {xyz[2]}"
                
                cv2.putText(dbg_result, result_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0,255,0) if found_intersect else (0,0,255), 2, cv2.LINE_AA)
                cv2.putText(dbg_result, coords_text, (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255,255,255), 2, cv2.LINE_AA)
                # ACTIVAR GUARDADO DE IMÁGENES DEBUG
                cv2.imwrite(f"debug_intersect_final_result_row{idx}.png", dbg_result)
                
                # Solo muestra debug textual detallado si NO encontró intersect en ninguna pasada
                if not found_intersect:
                    print(f"[DEBUG] Row {idx}: === TODAS LAS PASADAS FALLARON - ANÁLISIS DETALLADO ===")
                    print(f"[DEBUG] Row {idx}: Total detecciones fusionadas: {len(dets_merged)}")
                    for i, d in enumerate(dets_merged):
                        print(f"[DEBUG] Row {idx}: Fusionado {i}: '{d['text']}' (conf: {d['conf']:.1f})")
                    
                    if data_footer:
                        individual_high_conf = [
                            (i, data_footer["text"][i], data_footer["conf"][i]) 
                            for i in range(len(data_footer["text"])) 
                            if (data_footer["text"][i] or "").strip() and 
                               float(data_footer["conf"][i] if str(data_footer["conf"][i]).replace('.','',1).isdigit() else 0) >= 40
                        ]
                        print(f"[DEBUG] Row {idx}: Total detecciones individuales (conf>=40): {len(individual_high_conf)}")
                        for i, txt, conf in individual_high_conf[:10]:  # Solo primeras 10
                            print(f"[DEBUG] Row {idx}: Individual {i}: '{txt}' (conf: {conf})")

                # Asignacion final
                if xyz != ["NA", "NA", "NA"]:
                    X, Y, Z = xyz[0], xyz[1], xyz[2]
                    print(f"[DEBUG] Row {idx}: EXITO - X={X}, Y={Y}, Z={Z}")
                elif not found_intersect:
                    X, Y, Z = "NA", "NA", "NA"
                    print(f"[DEBUG] Row {idx}: NA porque no se encontro 'Intersect position'")
                elif not found_camera:
                    X, Y, Z = "NA", "NA", "NA"
                    print(f"[DEBUG] Row {idx}: NA porque no se encontro la camara")
                elif not found_station:
                    X, Y, Z = "NA", "NA", "NA"
                    print(f"[DEBUG] Row {idx}: NA porque no se encontro la estacion")
                else:
                    X, Y, Z = "NA", "NA", "NA"
                    print(f"[DEBUG] Row {idx}: NA porque no se extrajeron valores validos")

                try:
                    punto_nr = str(int(float(row[col_nr]))) if col_nr else str(idx+1)
                except:
                    punto_nr = str(idx+1)

                intersect_table.append([f"P {punto_nr}", X, Y, Z])
                print(f"[DEBUG] Intersect XYZ: {X}, {Y}, {Z}")
                print("\nTabla de puntos Intersect:")
                for row_print in intersect_table:
                    print(" ".join(row_print))

            
            # ESCRITURA DE COORDENADAS X, Y, Z EN EXCEL COMO NÚMERO (dentro del bucle)
            try:
                wb_write = openpyxl.load_workbook(target_excel)
                ws_write = wb_write[sheet_name]
                header_row = hdr + 1  # hdr viene de load_excel_any_header, openpyxl es 1-based

                # Buscar las columnas X_NEW, Y_NEW, Z_NEW (no crear si no existen)
                col_xnew = col_ynew = col_znew = None
                for col in range(1, ws_write.max_column + 1):
                    cell_val = ws_write.cell(row=header_row, column=col).value
                    if cell_val:
                        col_name = str(cell_val).strip().upper()
                        if col_name == "X_NEW":
                            col_xnew = col
                        elif col_name == "Y_NEW":
                            col_ynew = col
                        elif col_name == "Z_NEW":
                            col_znew = col
                if not all([col_xnew, col_ynew, col_znew]):
                    print(f"[ERROR] Alguna de las columnas X_NEW, Y_NEW o Z_NEW no existe en la hoja {sheet_name}. No se escribe nada.")
                else:
                    fila_punto = nr_to_row.get(str(punto_nr))
                    if not fila_punto:
                        print(f"[WARN] No se encontró la fila para el punto {punto_nr}, no se escribe nada.")
                    else:
                        # Guardar como número (float), usando coma o punto según formato
                        try:
                            x_num = float(X.replace(",", ".")) if isinstance(X, str) else float(X)
                            y_num = float(Y.replace(",", ".")) if isinstance(Y, str) else float(Y)
                            z_num = float(Z.replace(",", ".")) if isinstance(Z, str) else float(Z)
                            ws_write.cell(row=fila_punto, column=col_xnew).value = x_num
                            ws_write.cell(row=fila_punto, column=col_ynew).value = y_num
                            ws_write.cell(row=fila_punto, column=col_znew).value = z_num
                            print(f"[DEBUG] Escrito X={x_num}, Y={y_num}, Z={z_num} (numeros) en X_NEW, Y_NEW, Z_NEW para el punto {punto_nr} (fila {fila_punto})")
                        except Exception as conv_e:
                            ws_write.cell(row=fila_punto, column=col_xnew).value = X
                            ws_write.cell(row=fila_punto, column=col_ynew).value = Y
                            ws_write.cell(row=fila_punto, column=col_znew).value = Z
                            print(f"[WARN] No se pudo convertir alguna coordenada a número, guardadas como texto.")
                        wb_write.save(target_excel)
                wb_write.close()
            except Exception as e:
                print(f"[ERROR] No se pudo escribir en el Excel: {e}")

        except Exception as e:
            print(f"[ERR] Row {idx}: {e}")


if __name__ == "__main__":
    main()
